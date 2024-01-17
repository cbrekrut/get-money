from django.shortcuts import render,redirect,get_object_or_404
from django.http import HttpResponse
from django.http import JsonResponse
from openpyxl import Workbook
from django.views.decorators.http import require_POST
import json
from .models import Employee, Position, Task, Data
from .forms import UploadFileForm
import pandas as pd
from datetime import datetime,date
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F,Count

def daily_summary(request):
    print(request.GET)
    fio = request.GET.get('fio', '')
    position = request.GET.get('position', '')
    code = request.GET.get('task-code','')
    today = date.today()
    print(fio)
    print(position)
    print(code)
    # Фильтрация объектов Data на основе переданных параметров
    data_objects = Data.objects.filter(date=today, status='moderation', full_name=fio, position=position, code=code)
    print(data_objects)
    # Вычисление общей выручки
    total_earnings = data_objects.aggregate(total_earnings=Sum(F('count') * F('cost')))['total_earnings']
    
    # Получение выполненных задач и их количества
    completed_tasks = data_objects.values('task', 'count')

    return render(request, 'zavod/daily_summary.html', {'total_earnings': total_earnings, 'completed_tasks': completed_tasks,'fio':fio})

def employee_list(request):
    employees = Employee.objects.all()
    positions = Position.objects.all()
    tasks = Task.objects.all()
    task_codes = Task.objects.values_list('code', flat=True).distinct()

    return render(request, 'zavod/employee_list.html', {'employees': employees, 'positions': positions, 'tasks': tasks,'task_codes':task_codes})

def get_tasks(request):
    selected_position = request.GET.get('position', '')
    selected_task_code = request.GET.get('task_code', '')
    tasks = Task.objects.filter(position__title=selected_position, code=selected_task_code).values('code', 'name')
    return JsonResponse(list(tasks), safe=False)

def get_task_count(request):
    selected_task = request.GET.get('selected_task', '')
    selected_task_code = request.GET.get('task_code', '')
    print(selected_task_code)
    count_list = list(Task.objects.filter(name = selected_task,code = selected_task_code).values('count_detail'))
    data = {'count_list': count_list}
    print(data)
    return JsonResponse(data, safe=False)

def generate_excel(request):
    if request.method == 'POST':
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        data_objects = Data.objects.filter(date__range=(start_date, end_date),status='approved')
    else:
        data_objects = Data.objects.all()

    workbook = Workbook()
    worksheet = workbook.active

    headers = ['ФИО', 'Наименование цеха', 'Наименование операции','Код Заказа' ,'Дата', 'Стоимость', 'Кол-во', 'Финансы']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    for row_num, data_object in enumerate(data_objects, 2):
        worksheet.cell(row=row_num, column=1, value=data_object.full_name)
        worksheet.cell(row=row_num, column=2, value=data_object.position)
        worksheet.cell(row=row_num, column=3, value=data_object.task)
        worksheet.cell(row=row_num, column=4, value=data_object.code)
        worksheet.cell(row=row_num, column=5, value=data_object.date.strftime('%d-%m-%Y'))
        worksheet.cell(row=row_num, column=6, value=data_object.cost)
        worksheet.cell(row=row_num, column=7, value=data_object.count)
        worksheet.cell(row=row_num, column=8, value=(int(data_object.count) * float(data_object.cost)))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_data.xlsx'

    workbook.save(response)

    return response

@require_POST
def save_data(request):
    data = json.loads(request.body)
    fio = data.get('fio')
    position = data.get('position')
    tasks = data.get('tasks', [])
    code = data.get('code')
    tasks_count = data.get('tasks_count', [])
   
    for i, task_description in enumerate(tasks):
        task_count = int(tasks_count[i])

        task_instance = Task.objects.get(name=task_description,code =code)
        # Проверяем, чтобы count не стал отрицательным
        if task_instance.count_detail >= task_count:
            task_instance.count_detail -= task_count
            task_instance.save()
            data_instance = Data(full_name=fio, position=position, task=task_description, count=task_count,cost = task_instance.cost,code = task_instance.code)
            data_instance.save()
            return JsonResponse({'message': 'Data saved successfully'})
      
        

    return JsonResponse({'message': 'Data Failed'})

@login_required
def reports(request):
    employees = Employee.objects.all()
    return render(request,'zavod/reports.html',{'employees': employees})

@login_required
def orders(request):
    active_orders = Task.objects.filter(count_detail__gt=0)
    context = {'active_orders': active_orders}
    return render(request, 'zavod/orders.html', context)


@login_required
def upload_orders(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            df = pd.read_excel(file)

            for index, row in df.iterrows():
                code = row['Код заказа'] 
                name = row['Наименование операции']
                count_times = float(row['Кол-во операций/ Часов'])
                count_detail = int(row['Кол-во операций/ Деталей'])
                cost = row['Цена изготовления 1 детали']
                sels = int(row['Цена продажи 1 детали'])
                position_title = row['Цех участок']  

                position = get_object_or_404(Position, title=position_title)
                Task.objects.create(
                    code=code,
                    name=name,
                    count_times=count_times,
                    count_detail=count_detail,
                    cost=cost,
                    sels=sels,
                    position=position
                )

            return redirect('orders')  

    else:
        form = UploadFileForm()

    return render(request, 'zavod/upload_orders.html', {'form': form})

def workshop_report(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        workshop = request.POST.get('workshop')

        filtered_tasks = Data.objects.filter(
            date__range=[start_date, end_date],
            position=workshop
        )

        wb = Workbook()
        ws = wb.active

        headers = ['Цех', 'ФИО', 'Кол - во','Операция','Код заказа']
        ws.append(headers)

        for task in filtered_tasks:
            row_data = [task.position, task.full_name, task.count,task.task,task.code]
            ws.append(row_data)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=workshop_report_{start_date}_{end_date}.xlsx'
        wb.save(response)

    return response

def generate_employee_report_excel(employee, employee_data):
    wb = Workbook()
    ws = wb.active

    ws.append(['ФИО','Дата', 'Операция', 'Кол-во', 'Стоимость (закупка)'])
    for data in employee_data:
        ws.append([data.full_name, data.date, data.task, data.count, data.cost])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="employee_report.xlsx'

    wb.save(response)

    return response

def employee_report(request):
    if request.method == 'POST':
        employee = request.POST.get('employee')
        start_date_employee = request.POST.get('start_date_employee')
        end_date_employee = request.POST.get('end_date_employee')

        start_date_employee = datetime.strptime(start_date_employee, '%Y-%m-%d').date()
        end_date_employee = datetime.strptime(end_date_employee, '%Y-%m-%d').date()
        
        employee_data = Data.objects.filter(full_name=employee, date__range=(start_date_employee, end_date_employee))

        if not employee_data.exists():
            return HttpResponse("No data available for the selected employee and date range.")

        return generate_employee_report_excel(employee, employee_data)

    return HttpResponse("Invalid request method.")
